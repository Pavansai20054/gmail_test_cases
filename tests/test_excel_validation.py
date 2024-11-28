import openpyxl
import pytest
import os

@pytest.fixture
def load_excel_file():
    file_path = r"data\Pavan_Manual Testing Steps table.xlsx"
    if not os.path.exists(file_path):
        pytest.fail(f"File not found: {file_path}")
    workbook = openpyxl.load_workbook(file_path)
    return workbook

def test_sheet_exists(load_excel_file):
    # Check if the required sheet exists
    workbook = load_excel_file
    assert "Sheet1" in workbook.sheetnames, "Sheet1 is missing in the workbook"

def test_column_names(load_excel_file):
    # Check if column names are as expected
    workbook = load_excel_file
    sheet = workbook["Sheet1"]
    expected_columns = ["Step", "Description", "Expected Result"]
    actual_columns = [sheet.cell(row=1, column=col).value for col in range(1, len(expected_columns) + 1)]
    assert actual_columns == expected_columns, \
        f"Expected columns {expected_columns}, but found {actual_columns}"

def test_no_empty_cells(load_excel_file):
    # Check for empty cells in critical columns
    workbook = load_excel_file
    sheet = workbook["Sheet1"]
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
        assert all(cell.value is not None for cell in row), \
            f"Empty cell found in row {row[0].row}"
